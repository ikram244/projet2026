"""
api_mesures_route.py
----------------------
Blueprint Flask : saisie et consultation des mesures de laboratoire.

POST /api/mesures                    -> creer une mesure (reserve au role TECHNICIEN_LABO)
GET  /api/mesures                    -> liste des mesures (accessible a tous les roles connectes)
GET  /api/densite-cible/derniere     -> derniere densite cible saisie en page Correction
GET  /api/mesures/export             -> export Excel des mesures filtrees
"""
from io import BytesIO

from flask import Blueprint, request, jsonify, session, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from db import (
    create_mesure, get_mesures, get_derniere_densite_cible,
    create_alertes_pour_tous_sauf_stagiaire, get_mesures_filtrees,
)

mesures_bp = Blueprint("mesures_bp", __name__)


@mesures_bp.route("/api/mesures", methods=["POST"])
def add_mesure():
    if "user_id" not in session:
        return jsonify({"error": "Non authentifié."}), 401
    if session.get("user_role") != "TECHNICIEN_LABO":
        return jsonify({"error": "Seul un technicien laboratoire peut saisir une mesure."}), 403

    data = request.get_json(force=True)

    echelon = data.get("echelon")
    date_prelevement = data.get("date_prelevement")
    heure_prelevement = data.get("heure_prelevement")
    densite_entree_29 = data.get("densite_entree_29")
    densite_sortie_54 = data.get("densite_sortie_54")
    etat_echelon = data.get("etat_echelon", "FONCTIONNEL")

    if etat_echelon not in ("FONCTIONNEL", "ARRET"):
        etat_echelon = "FONCTIONNEL"

    if not echelon or not date_prelevement or not heure_prelevement:
        return jsonify({"error": "Échelon, date et heure de prélèvement sont requis."}), 400

    new_id = create_mesure(
        session["user_id"], echelon, date_prelevement, heure_prelevement,
        densite_entree_29, densite_sortie_54, etat_echelon,
    )
    if not new_id:
        return jsonify({"error": "Erreur lors de l'enregistrement."}), 500

    d29 = densite_entree_29 if densite_entree_29 is not None else "—"
    d54 = densite_sortie_54 if densite_sortie_54 is not None else "—"
    message = f"Densité mesurée — Échelon {echelon} : ACP29% = {d29}, ACP54% = {d54}"
    create_alertes_pour_tous_sauf_stagiaire(message, exclure_utilisateur_id=session["user_id"])

    return jsonify({"id": new_id}), 201


@mesures_bp.route("/api/mesures", methods=["GET"])
def list_mesures():
    if "user_id" not in session:
        return jsonify({"error": "Non authentifié."}), 401

    rows = get_mesures()
    for r in rows or []:
        if r.get("date_prelevement"):
            r["date_prelevement"] = r["date_prelevement"].strftime("%Y-%m-%d")
        if r.get("heure_prelevement"):
            r["heure_prelevement"] = str(r["heure_prelevement"])
        if r.get("date_saisie"):
            r["date_saisie"] = r["date_saisie"].strftime("%Y-%m-%d %H:%M")

    return jsonify(rows or [])


@mesures_bp.route("/api/densite-cible/derniere", methods=["GET"])
def derniere_densite_cible():
    if "user_id" not in session:
        return jsonify({"error": "Non authentifié."}), 401

    echelon = request.args.get("echelon")
    row = get_derniere_densite_cible(echelon)
    if not row:
        return jsonify({"densite_cible": None})

    return jsonify({"densite_cible": row["densite_cible"]})


@mesures_bp.route("/api/mesures/export", methods=["GET"])
def export_mesures():
    if "user_id" not in session:
        return jsonify({"error": "Non authentifié."}), 401

    date_debut = request.args.get("date_debut")
    date_fin = request.args.get("date_fin")
    echelon = request.args.get("echelon", "all")

    rows = get_mesures_filtrees(date_debut, date_fin, echelon) or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Historique mesures"

    headers = ["Date", "Heure", "Échelon", "État", "Densité ACP 29%", "Densité ACP 54%", "Saisi par"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1A6B3A", end_color="1A6B3A", fill_type="solid")

    for r in rows:
        ws.append([
            r["date_prelevement"].strftime("%Y-%m-%d") if r.get("date_prelevement") else "",
            str(r["heure_prelevement"]) if r.get("heure_prelevement") else "",
            r.get("echelon", ""),
            r.get("etat_echelon", ""),
            float(r["densite_entree_29"]) if r.get("densite_entree_29") is not None else None,
            float(r["densite_sortie_54"]) if r.get("densite_sortie_54") is not None else None,
            f"{r.get('prenom', '')} {r.get('nom', '')}",
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 3

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="historique_mesures_labo.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )